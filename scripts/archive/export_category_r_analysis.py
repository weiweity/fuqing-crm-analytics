"""
芙清CRM - 品类30指标对比 + R区间 导出脚本
时间范围: 2025-05-06 ~ 2025-06-21
品类: 凉茶次抛 / 经典膜 / 凉茶水乳
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from backend.db.connection import get_connection
from backend.services.category_service import EXCLUDED_PRODUCT_CATEGORIES
from backend.semantic.filters import OrderFilters

START_DATE = "2026-05-06"
END_DATE = "2026-06-21"
CUTOFF_DT = "2026-05-05"
CATEGORIES = ["凉茶次抛", "经典膜", "凉茶水乳"]

CAT_EXPR = "COALESCE(TRIM(o.spu_product_subclass), '未知')"
VALID_SQL, _valid_params = OrderFilters.valid_order()  # 语义层统一口径
EXCL_PLACEHOLDERS = ",".join(["?"] * len(EXCLUDED_PRODUCT_CATEGORIES))
EXCL_SQL = f"AND {CAT_EXPR} NOT IN ({EXCL_PLACEHOLDERS})"
EXCL_PARAMS = list(EXCLUDED_PRODUCT_CATEGORIES)


def _run(conn, sql, params):
    return conn.execute(sql, params).fetchone()


# ============================================================
# 1. 30指标对比
# ============================================================
def compute_30_metrics(conn, start_date, end_date):
    rows = []
    for cat in CATEGORIES:
        # params: [is_new_start, pay_start, pay_end] + 18个excl + [cat]
        params = [start_date, start_date, end_date] + EXCL_PARAMS + [cat]

        sql = f"""
        WITH period_orders AS (
            SELECT
                {CAT_EXPR} AS cat_name,
                o.user_id,
                o.actual_amount,
                o.is_member,
                CASE WHEN ufp.first_pay_date >= DATE(?) THEN 1 ELSE 0 END AS is_new
            FROM orders o
            LEFT JOIN user_first_purchase ufp ON o.user_id = ufp.user_id
            WHERE o.pay_time >= ?
              AND o.pay_time < DATE(?) + INTERVAL '1' DAY
              AND {VALID_SQL}
              {EXCL_SQL}
        )
        SELECT
            COUNT(DISTINCT user_id) AS total_users,
            SUM(actual_amount) AS total_gsv,
            SUM(CASE WHEN is_member THEN actual_amount ELSE 0 END) AS member_gsv,
            COUNT(DISTINCT CASE WHEN is_member THEN user_id END) AS member_users,
            SUM(CASE WHEN is_new = 0 THEN actual_amount ELSE 0 END) AS old_gsv,
            COUNT(DISTINCT CASE WHEN is_new = 0 THEN user_id END) AS old_users,
            SUM(CASE WHEN is_new = 1 THEN actual_amount ELSE 0 END) AS new_gsv,
            COUNT(DISTINCT CASE WHEN is_new = 1 THEN user_id END) AS new_users
        FROM period_orders
        WHERE cat_name = ?
        """
        r = _run(conn, sql, params)
        total_users = int(r[0] or 0)
        total_gsv = float(r[1] or 0)
        member_gsv = float(r[2] or 0)
        member_users = int(r[3] or 0)
        old_gsv = float(r[4] or 0)
        old_users = int(r[5] or 0)
        new_gsv = float(r[6] or 0)
        new_users = int(r[7] or 0)

        rows.append({
            "品类": cat,
            "GSV": round(total_gsv, 2),
            "GSV_万": round(total_gsv / 10000, 1),
            "人数": total_users,
            "AUS": round(total_gsv / total_users, 2) if total_users > 0 else 0,
            "会员GSV": round(member_gsv, 2),
            "会员占比": round(member_gsv / total_gsv, 4) if total_gsv > 0 else 0,
            "会员人数": member_users,
            "老客GSV": round(old_gsv, 2),
            "老客GSV占比": round(old_gsv / total_gsv, 4) if total_gsv > 0 else 0,
            "老客人数": old_users,
            "老客AUS": round(old_gsv / old_users, 2) if old_users > 0 else 0,
            "新客GSV": round(new_gsv, 2),
            "新客GSV占比": round(new_gsv / total_gsv, 4) if total_gsv > 0 else 0,
            "新客人数": new_users,
            "新客AUS": round(new_gsv / new_users, 2) if new_users > 0 else 0,
        })
    return rows


# ============================================================
# 2. R区间品类人群版
# ============================================================
R_BUCKETS = [
    ("近1月", 0, 30),
    ("2-3月", 31, 90),
    ("4-6月", 91, 180),
    ("7-12月", 181, 365),
    ("13-24月", 366, 730),
    ("2年外", 731, 99999),
    ("TTL", 0, 99999),
]


def compute_r区间_cat_version(conn, cutoff_dt, start_date, end_date):
    results = {}
    for cat in CATEGORIES:
        cat_rows = []
        for label, r_min, r_max in R_BUCKETS:
            if label == "TTL":
                # 历史所有该品类用户（截至cutoff_dt）
                hist_sql = f"""
                SELECT COUNT(DISTINCT o.user_id)
                FROM orders o
                WHERE o.pay_time <= DATE(?) + INTERVAL '1' DAY
                  AND o.pay_time >= '2000-01-01'
                  AND {VALID_SQL}
                  {EXCL_SQL}
                  AND {CAT_EXPR} = ?
                """
                hist_count = _run(conn, hist_sql, [cutoff_dt] + EXCL_PARAMS + [cat])[0] or 0

                # 回购：分析期购买该品类的历史用户
                rep_sql = f"""
                WITH hist_users AS (
                    SELECT DISTINCT o.user_id
                    FROM orders o
                    WHERE o.pay_time <= DATE(?) + INTERVAL '1' DAY
                      AND o.pay_time >= '2000-01-01'
                      AND {VALID_SQL}
                      {EXCL_SQL}
                      AND {CAT_EXPR} = ?
                ),
                rep_users AS (
                    SELECT DISTINCT o.user_id
                    FROM orders o
                    INNER JOIN hist_users h ON o.user_id = h.user_id
                    WHERE o.pay_time >= ?
                      AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                      AND {VALID_SQL}
                      AND {CAT_EXPR} = ?
                )
                SELECT COUNT(DISTINCT o.user_id), COALESCE(SUM(o.actual_amount), 0)
                FROM orders o
                WHERE o.user_id IN (SELECT user_id FROM rep_users)
                  AND o.pay_time >= ?
                  AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                  AND {VALID_SQL}
                  AND {CAT_EXPR} = ?
                """
                rep_r = _run(conn, rep_sql, [cutoff_dt] + EXCL_PARAMS + [cat] + [start_date, end_date, cat] + [start_date, end_date, cat])
                # 特定R区间：历史用户中，距cutoff_dt在[r_min, r_max]天内的
                hist_sql = f"""
                WITH base AS (
                    SELECT o.user_id, MAX(o.pay_time) AS last_pay
                    FROM orders o
                    WHERE o.pay_time <= DATE(?) + INTERVAL '1' DAY
                      AND o.pay_time >= '2000-01-01'
                      AND {VALID_SQL}
                      {EXCL_SQL}
                      AND {CAT_EXPR} = ?
                    GROUP BY o.user_id
                    HAVING DATEDIFF('day', MAX(o.pay_time), DATE(?)) BETWEEN ? AND ?
                )
                SELECT COUNT(*) FROM base
                """
                hist_count = _run(conn, hist_sql, [cutoff_dt] + EXCL_PARAMS + [cat] + [cutoff_dt, r_min, r_max])[0] or 0

                rep_sql = f"""
                WITH base AS (
                    SELECT o.user_id, MAX(o.pay_time) AS last_pay
                    FROM orders o
                    WHERE o.pay_time <= DATE(?) + INTERVAL '1' DAY
                      AND o.pay_time >= '2000-01-01'
                      AND {VALID_SQL}
                      {EXCL_SQL}
                      AND {CAT_EXPR} = ?
                    GROUP BY o.user_id
                    HAVING DATEDIFF('day', MAX(o.pay_time), DATE(?)) BETWEEN ? AND ?
                ),
                rep_users AS (
                    SELECT DISTINCT o.user_id
                    FROM orders o
                    INNER JOIN base b ON o.user_id = b.user_id
                    WHERE o.pay_time >= ?
                      AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                      AND {VALID_SQL}
                      AND {CAT_EXPR} = ?
                )
                SELECT COUNT(DISTINCT o.user_id), COALESCE(SUM(o.actual_amount), 0)
                FROM orders o
                WHERE o.user_id IN (SELECT user_id FROM rep_users)
                  AND o.pay_time >= ?
                  AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                  AND {VALID_SQL}
                  AND {CAT_EXPR} = ?
                """
                rep_r = _run(conn, rep_sql, [cutoff_dt] + EXCL_PARAMS + [cat] + [cutoff_dt, r_min, r_max] + [start_date, end_date, cat] + [start_date, end_date, cat])
                rep_count = int(rep_r[0] or 0)
                rep_gsv = float(rep_r[1] or 0)
                rate = round(rep_count / hist_count, 4) if hist_count > 0 else 0
                cat_rows.append({"R区间": label, "历史人数": int(hist_count), "回购人数": rep_count, "回购率": rate, "回购GSV": round(rep_gsv, 2)})

        results[cat] = cat_rows
    return results


# ============================================================
# 3. R区间全店人群版
# ============================================================
def compute_r区间_全店版本(conn, cutoff_dt, start_date, end_date):
    results = {}
    for cat in CATEGORIES:
        cat_rows = []
        for label, r_min, r_max in R_BUCKETS:
            if label == "TTL":
                # 全店历史所有用户（无品类过滤）
                hist_sql = """
                SELECT COUNT(DISTINCT user_id)
                FROM orders
                WHERE pay_time <= DATE(?) + INTERVAL '1' DAY
                  AND pay_time >= '2000-01-01'
                  AND is_goujinjin = FALSE AND order_status != '交易关闭' AND is_refund = FALSE
                """
                hist_count = _run(conn, hist_sql, [cutoff_dt])[0] or 0

                # 回购：分析期购买该品类的全店历史用户
                rep_sql = f"""
                WITH rep_users AS (
                    SELECT DISTINCT o.user_id
                    FROM orders o
                    WHERE o.pay_time >= ?
                      AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                      AND {VALID_SQL}
                      AND {CAT_EXPR} = ?
                )
                SELECT COUNT(DISTINCT o.user_id), COALESCE(SUM(o.actual_amount), 0)
                FROM orders o
                WHERE o.user_id IN (SELECT user_id FROM rep_users)
                  AND o.pay_time >= ?
                  AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                  AND {VALID_SQL}
                  AND {CAT_EXPR} = ?
                """
                rep_r = _run(conn, rep_sql, [start_date, end_date, cat] + [start_date, end_date, cat])
                rep_count = int(rep_r[0] or 0)
                rep_gsv = float(rep_r[1] or 0)
                rate = round(rep_count / hist_count, 4) if hist_count > 0 else 0
                cat_rows.append({"R区间": label, "全店历史人数": int(hist_count), "本品类回购人数": rep_count, "回购率": rate, "本品类回购GSV": round(rep_gsv, 2)})
            else:
                # 特定R区间
                hist_sql = """
                WITH base AS (
                    SELECT user_id, MAX(pay_time) AS last_pay
                    FROM orders
                    WHERE pay_time <= DATE(?) + INTERVAL '1' DAY
                      AND pay_time >= '2000-01-01'
                      AND is_goujinjin = FALSE AND order_status != '交易关闭' AND is_refund = FALSE
                    GROUP BY user_id
                    HAVING DATEDIFF('day', MAX(pay_time), DATE(?)) BETWEEN ? AND ?
                )
                SELECT COUNT(*) FROM base
                """
                hist_count = _run(conn, hist_sql, [cutoff_dt, cutoff_dt, r_min, r_max])[0] or 0

                rep_sql = f"""
                WITH base AS (
                    SELECT user_id, MAX(pay_time) AS last_pay
                    FROM orders
                    WHERE pay_time <= DATE(?) + INTERVAL '1' DAY
                      AND pay_time >= '2000-01-01'
                      AND is_goujinjin = FALSE AND order_status != '交易关闭' AND is_refund = FALSE
                    GROUP BY user_id
                    HAVING DATEDIFF('day', MAX(pay_time), DATE(?)) BETWEEN ? AND ?
                ),
                rep_users AS (
                    SELECT DISTINCT o.user_id
                    FROM orders o
                    INNER JOIN base b ON o.user_id = b.user_id
                    WHERE o.pay_time >= ?
                      AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                      AND {VALID_SQL}
                      AND {CAT_EXPR} = ?
                )
                SELECT COUNT(DISTINCT o.user_id), COALESCE(SUM(o.actual_amount), 0)
                FROM orders o
                WHERE o.user_id IN (SELECT user_id FROM rep_users)
                  AND o.pay_time >= ?
                  AND o.pay_time < DATE(?) + INTERVAL '1' DAY
                  AND {VALID_SQL}
                  AND {CAT_EXPR} = ?
                """
                rep_r = _run(conn, rep_sql, [cutoff_dt, cutoff_dt, r_min, r_max] + [start_date, end_date, cat] + [start_date, end_date, cat])
                rep_count = int(rep_r[0] or 0)
                rep_gsv = float(rep_r[1] or 0)
                rate = round(rep_count / hist_count, 4) if hist_count > 0 else 0
                cat_rows.append({"R区间": label, "全店历史人数": int(hist_count), "本品类回购人数": rep_count, "回购率": rate, "本品类回购GSV": round(rep_gsv, 2)})

        results[cat] = cat_rows
    return results


# ============================================================
# Excel 导出
# ============================================================
def style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center", vertical="center")


def export_30_metrics(rows, out_path):
    headers = ["品类", "GSV", "GSV_万", "人数", "AUS",
               "会员GSV", "会员占比", "会员人数",
               "老客GSV", "老客GSV占比", "老客人数", "老客AUS",
               "新客GSV", "新客GSV占比", "新客人数", "新客AUS"]
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    style_header(ws, headers)
    for r in rows:
        ws.append([r[h] for h in headers])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    wb.save(out_path)
    print(f"已导出: {out_path}")


def export_r区间_cat(rows_by_cat, out_path):
    wb = Workbook()
    for cat in CATEGORIES:
        ws = wb.create_sheet(cat)
        headers = ["R区间", "历史人数", "回购人数", "回购率", "回购GSV"]
        ws.append(headers)
        style_header(ws, headers)
        for r in rows_by_cat[cat]:
            ws.append([r[h] for h in headers])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
    wb.remove(wb.active)
    wb.save(out_path)
    print(f"已导出: {out_path}")


def export_r区间_全店(rows_by_cat, out_path):
    wb = Workbook()
    for cat in CATEGORIES:
        ws = wb.create_sheet(cat)
        headers = ["R区间", "全店历史人数", "本品类回购人数", "回购率", "本品类回购GSV"]
        ws.append(headers)
        style_header(ws, headers)
        for r in rows_by_cat[cat]:
            ws.append([r[h] for h in headers])
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    wb.remove(wb.active)
    wb.save(out_path)
    print(f"已导出: {out_path}")


if __name__ == "__main__":
    conn = get_connection()
    try:
        print("计算30指标对比...")
        rows = compute_30_metrics(conn, START_DATE, END_DATE)
        for r in rows:
            print(f"  {r['品类']}: GSV={r['GSV_万']}万 人数={r['人数']}")

        print("\n计算R区间（品类人群版）...")
        r_cat = compute_r区间_cat_version(conn, CUTOFF_DT, START_DATE, END_DATE)

        print("\n计算R区间（全店人群版）...")
        r_all = compute_r区间_全店版本(conn, CUTOFF_DT, START_DATE, END_DATE)

        base = "/Users/hutou/Desktop/fuqin date/sample-crm-analytics/exports"
        os.makedirs(base, exist_ok=True)

        export_30_metrics(rows, f"{base}/品类30指标对比_2026年5月6日-6月21日.xlsx")
        export_r区间_cat(r_cat, f"{base}/品类R区间_2026年5月6日-6月21日.xlsx")
        export_r区间_全店(r_all, f"{base}/品类R区间_全店人群_2026年5月6日-6月21日.xlsx")
        print("\n全部完成!")
    finally:
        conn.close()
