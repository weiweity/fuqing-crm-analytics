"""Sprint 171 ad-hoc-query XLSX style SSOT.

Excel 视觉规范:
- 表头深蓝 #1F4E79, 子标题中蓝 #2E75B6
- 同比正值 A 股红 #D32F2F, 负值绿色 #2E7D32
- 0 公式: 写入前拒绝以 "=" 开头的字符串
"""
from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THEME_HEADER_HEX = "1F4E79"
THEME_SUBHEADER_HEX = "2E75B6"
YOY_POS_HEX = "D32F2F"
YOY_NEG_HEX = "2E7D32"

THEME_HEADER = PatternFill(start_color=THEME_HEADER_HEX, end_color=THEME_HEADER_HEX, fill_type="solid")
THEME_SUBHEADER = PatternFill(start_color=THEME_SUBHEADER_HEX, end_color=THEME_SUBHEADER_HEX, fill_type="solid")
FONT_HEADER = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Microsoft YaHei", size=10)
FONT_YOY_POS = Font(name="Microsoft YaHei", size=10, bold=True, color=YOY_POS_HEX)
FONT_YOY_NEG = Font(name="Microsoft YaHei", size=10, bold=True, color=YOY_NEG_HEX)
FONT_TITLE = Font(name="Microsoft YaHei", size=13, bold=True, color=THEME_HEADER_HEX)

THIN_SIDE = Side(style="thin", color="D9E2F3")
DEFAULT_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _assert_not_formula(value: Any) -> None:
    if isinstance(value, str) and value.startswith("="):
        raise ValueError("XLSX output forbids formulas")


def apply_header(cell) -> None:
    cell.fill = THEME_HEADER
    cell.font = FONT_HEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = DEFAULT_BORDER


def apply_subheader(cell) -> None:
    cell.fill = THEME_SUBHEADER
    cell.font = FONT_SUBHEADER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = DEFAULT_BORDER


def apply_body(cell) -> None:
    cell.font = FONT_BODY
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = DEFAULT_BORDER


def apply_yoy(cell, value: Any, unit: str = "%") -> None:
    """写同比单元格并按正负着色。value 是已 *100 的百分比或 pp。"""
    _assert_not_formula(value)
    if value is None or value == "":
        cell.value = "N/A"
        apply_body(cell)
        return
    number = float(value)
    cell.font = FONT_YOY_POS if number >= 0 else FONT_YOY_NEG
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = DEFAULT_BORDER
    suffix = "pp" if unit == "pp" else "%"
    cell.value = f"+{number:.2f}{suffix}" if number >= 0 else f"{number:.2f}{suffix}"


def normalize_sheet_title(title: str) -> str:
    return title[:31]


def write_rows_to_sheet(
    ws,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    title: str | None = None,
    yoy_columns: set[int] | None = None,
    pp_columns: set[int] | None = None,
) -> None:
    """写一个标准表格到 worksheet，列索引从 0 开始。"""
    current_row = 1
    if title:
        ws.cell(row=current_row, column=1, value=title)
        ws.cell(row=current_row, column=1).font = FONT_TITLE
        current_row += 1

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        apply_header(cell)
    header_row = current_row

    yoy_columns = yoy_columns or set()
    pp_columns = pp_columns or set()
    current_row += 1
    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            _assert_not_formula(value)
            zero_idx = col_idx - 1
            cell = ws.cell(row=current_row, column=col_idx)
            if zero_idx in yoy_columns:
                apply_yoy(cell, value, unit="pp" if zero_idx in pp_columns else "%")
            else:
                cell.value = value
                apply_body(cell)
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        current_row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 34)


def reserve_output_path(output_path: str | None, default_name: str = "ad-hoc-query.xlsx") -> Path:
    """用 O_EXCL 预留输出路径，避免同秒覆盖。"""
    out = Path(output_path) if output_path else Path("/tmp") / default_name
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.suffix.lower() != ".xlsx":
        out = out.with_suffix(".xlsx")
    while True:
        candidate = out
        if candidate.exists():
            suffix = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            candidate = out.with_name(f"{out.stem}_{suffix}{out.suffix}")
        try:
            fd = os.open(str(candidate), os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o644)
            os.close(fd)
            return candidate
        except FileExistsError:
            out = out.with_name(f"{out.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{out.suffix}")


def save_workbook(workbook: Workbook, output_path: str | None, default_name: str = "ad-hoc-query.xlsx") -> str:
    path = reserve_output_path(output_path, default_name=default_name)
    workbook.save(path)
    return str(path)


def write_table_workbook(
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    output_path: str | None,
    sheet_name: str,
    title: str | None = None,
) -> str:
    workbook = Workbook()
    ws = workbook.active
    ws.title = normalize_sheet_title(sheet_name)
    write_rows_to_sheet(ws, headers=headers, rows=rows, title=title)
    return save_workbook(workbook, output_path)
