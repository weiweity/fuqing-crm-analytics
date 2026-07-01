#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按商品ID查询三个时间段新老客指标，并新增妆品/械品/淘客品 TTL 汇总行。
输出表头：链接归类、单品归类、商品ID + 各周期/同比/环比下老客/老客GSV/老客客单/新客/新客GSV/新客客单
"""
from __future__ import annotations

import csv
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.ad_hoc_queries._utils import read_only_conn, build_take_path, _sanitize_path_component

# fmt: off
PRODUCT_IDS = [
    # 妆品
    "803474428381", "870597889980", "683395365107", "660012488593", "803417397714",
    "654390297284", "753548858886", "674944410250", "900975734816", "800903824350",
    "656561260141", "912388775061", "933524395698", "781713844237", "1010458880710",
    "1009707365820", "680535358203", "831602598308", "967971403905", "978994528428",
    # 械品
    "597655781410", "587051744204", "587554886491", "601760206476", "587053192746",
    "612503357090", "781706928918", "871040351635", "684155734133", "836974872996",
    "627423052420", "847821870403",
    # 淘客品
    "621639424901", "773698929360", "789483733628",
]

CATEGORY_GROUPS = {
    "妆品销售TTL": [
        "803474428381", "870597889980", "683395365107", "660012488593", "803417397714",
        "654390297284", "753548858886", "674944410250", "900975734816", "800903824350",
        "656561260141", "912388775061", "933524395698", "781713844237", "1010458880710",
        "1009707365820", "680535358203", "831602598308", "967971403905", "978994528428",
    ],
    "械品销售TTL": [
        "597655781410", "587051744204", "587554886491", "601760206476", "587053192746",
        "612503357090", "781706928918", "871040351635", "684155734133", "836974872996",
        "627423052420", "847821870403",
    ],
    "淘客品销售TTL": [
        "621639424901", "773698929360", "789483733628",
    ],
}
# fmt: on

PERIODS = {
    "Y26_0506_0621": ("2026-05-06", "2026-06-21"),
    "Y25_0506_0621": ("2025-05-06", "2025-06-21"),
    "Y25_0929_1114": ("2025-09-29", "2025-11-14"),
}


def _build_id_values(ids: list[str]) -> str:
    return ",".join(f"'{pid}'" for pid in ids)


def _build_base_sql(ids: list[str]) -> str:
    id_values = _build_id_values(ids)
    return f"""
WITH base AS (
    SELECT
        o.product_id,
        o.product_title,
        o.spu_category,
        o.spu_product_class,
        o.user_id,
        o.actual_amount,
        o.is_refund,
        o.is_goujinjin,
        o.order_status,
        o.pay_time
    FROM orders o
    WHERE o.product_id IN ({id_values})
      AND o.is_goujinjin = FALSE
      AND o.order_status != '交易关闭'
      AND o.pay_time >= ?::TIMESTAMP AND o.pay_time < ?::TIMESTAMP
),
old_flag AS (
    SELECT
        b.*,
        CASE WHEN ufp.first_pay_date <= (CAST(b.pay_time AS DATE) - INTERVAL 1 DAY)
             THEN 1 ELSE 0 END AS is_old
    FROM base b
    LEFT JOIN user_first_purchase ufp ON b.user_id = ufp.user_id
)
"""


def _build_product_sql(ids: list[str]) -> str:
    return _build_base_sql(ids) + """
SELECT
    product_id,
    MAX(product_title) AS product_title,
    MAX(spu_category) AS spu_category,
    MAX(spu_product_class) AS spu_product_class,
    COUNT(DISTINCT CASE WHEN is_old = 1 AND is_refund = FALSE THEN user_id END) AS old_users,
    COALESCE(SUM(CASE WHEN is_old = 1 AND is_refund = FALSE THEN actual_amount ELSE 0 END), 0) AS old_gsv,
    COUNT(DISTINCT CASE WHEN is_old = 0 AND is_refund = FALSE THEN user_id END) AS new_users,
    COALESCE(SUM(CASE WHEN is_old = 0 AND is_refund = FALSE THEN actual_amount ELSE 0 END), 0) AS new_gsv
FROM old_flag
GROUP BY product_id
"""


def _build_group_sql(ids: list[str]) -> str:
    return _build_base_sql(ids) + """
SELECT
    COUNT(DISTINCT CASE WHEN is_old = 1 AND is_refund = FALSE THEN user_id END) AS old_users,
    COALESCE(SUM(CASE WHEN is_old = 1 AND is_refund = FALSE THEN actual_amount ELSE 0 END), 0) AS old_gsv,
    COUNT(DISTINCT CASE WHEN is_old = 0 AND is_refund = FALSE THEN user_id END) AS new_users,
    COALESCE(SUM(CASE WHEN is_old = 0 AND is_refund = FALSE THEN actual_amount ELSE 0 END), 0) AS new_gsv
FROM old_flag
"""


def _period_key(start: str, end: str) -> str:
    for k, (s, e) in PERIODS.items():
        if s == start and e == end:
            return k
    return f"{start}_{end}"


def _safe_yoy(cur: Optional[float], comp: Optional[float]) -> Optional[float]:
    if cur is None or comp is None:
        return None
    if comp == 0:
        return None if cur == 0 else None
    return round((cur - comp) / comp * 100, 2)


def _row_from_record(r: tuple) -> dict:
    """从查询结果 tuple 构造指标字典。"""
    old_users = int(r[0]) if r[0] else 0
    old_gsv = float(r[1]) if r[1] is not None else 0.0
    new_users = int(r[2]) if r[2] else 0
    new_gsv = float(r[3]) if r[3] is not None else 0.0
    return {
        "old_users": old_users,
        "old_gsv": old_gsv,
        "old_aus": round(old_gsv / old_users, 2) if old_users > 0 else 0.0,
        "new_users": new_users,
        "new_gsv": new_gsv,
        "new_aus": round(new_gsv / new_users, 2) if new_users > 0 else 0.0,
    }


def _run_query(ids: list[str], start: str, end: str) -> dict[str, dict]:
    """按商品ID返回每个商品的新老客指标。"""
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end, "%Y-%m-%d").date()
    start_dt = f"{s} 00:00:00"
    end_excl_dt = f"{e + timedelta(days=1)} 00:00:00"

    sql = _build_product_sql(ids)
    params = [start_dt, end_excl_dt]

    result: dict[str, dict] = {}
    with read_only_conn() as conn:
        cur = conn.execute(sql, params)
        for r in cur.fetchall():
            pid = str(r[0])
            metrics = _row_from_record(r[4:])
            result[pid] = {
                "product_id": pid,
                "product_title": r[1] or "",
                "spu_category": r[2] or "",
                "spu_product_class": r[3] or "",
                **metrics,
            }
    return result


def _run_group_query(ids: list[str], start: str, end: str) -> dict:
    """按商品ID列表聚合，返回该分类的新老客指标。"""
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end, "%Y-%m-%d").date()
    start_dt = f"{s} 00:00:00"
    end_excl_dt = f"{e + timedelta(days=1)} 00:00:00"

    sql = _build_group_sql(ids)
    params = [start_dt, end_excl_dt]

    with read_only_conn() as conn:
        cur = conn.execute(sql, params)
        r = cur.fetchone()
        if r is None:
            return _row_from_record((0, 0, 0, 0))
        return _row_from_record(r)


def _make_header() -> list[str]:
    header = ["链接归类", "单品归类", "商品ID"]
    sections = [
        ("Y26 5/6-6/21", ["Y26_0506_0621"]),
        ("同比", ["Y26_0506_0621", "Y25_0506_0621"]),
        ("环比", ["Y26_0506_0621", "Y25_0929_1114"]),
        ("Y25 5/6-6/21", ["Y25_0506_0621"]),
        ("Y25 9/29-11/14", ["Y25_0929_1114"]),
    ]
    for section_name, _ in sections:
        for metric in ["老客", "老客GSV", "老客客单", "新客", "新客GSV", "新客客单"]:
            header.append(f"{section_name}_{metric}")
    return header


def _build_row(spu_category: str, spu_product_class: str, product_label: str,
               metric_source: dict) -> list:
    base = metric_source.get("Y26_0506_0621", {})
    y25_same = metric_source.get("Y25_0506_0621", {})
    y25_prev = metric_source.get("Y25_0929_1114", {})

    def get(base_data, k):
        return base_data.get(k, 0) if base_data else 0

    def yoy_row(field: str) -> Optional[float]:
        return _safe_yoy(get(base, field), get(y25_same, field))

    def pop_row(field: str) -> Optional[float]:
        return _safe_yoy(get(base, field), get(y25_prev, field))

    return [
        spu_category,
        spu_product_class,
        product_label,
        # Y26 5/6-6/21
        get(base, "old_users"), get(base, "old_gsv"), get(base, "old_aus"),
        get(base, "new_users"), get(base, "new_gsv"), get(base, "new_aus"),
        # 同比 (Y26 vs Y25 5/6-6/21)
        yoy_row("old_users"), yoy_row("old_gsv"), yoy_row("old_aus"),
        yoy_row("new_users"), yoy_row("new_gsv"), yoy_row("new_aus"),
        # 环比 (Y26 vs Y25 9/29-11/14)
        pop_row("old_users"), pop_row("old_gsv"), pop_row("old_aus"),
        pop_row("new_users"), pop_row("new_gsv"), pop_row("new_aus"),
        # Y25 5/6-6/21
        get(y25_same, "old_users"), get(y25_same, "old_gsv"), get(y25_same, "old_aus"),
        get(y25_same, "new_users"), get(y25_same, "new_gsv"), get(y25_same, "new_aus"),
        # Y25 9/29-11/14
        get(y25_prev, "old_users"), get(y25_prev, "old_gsv"), get(y25_prev, "old_aus"),
        get(y25_prev, "new_users"), get(y25_prev, "new_gsv"), get(y25_prev, "new_aus"),
    ]


def _write_excel(rows: list[list], summary_rows: list[int], output_path: Path) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[WARN] openpyxl not available, skip Excel", file=sys.stderr)
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品新老客对比"

    header = _make_header()
    ws.append(header)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    align = Alignment(horizontal="center", vertical="center")
    summary_fill = PatternFill("solid", fgColor="D9E1F2")
    summary_font = Font(bold=True)
    category_title_font = Font(bold=True, color="1F4E79")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = thin_border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for i, cell in enumerate(row):
            cell.border = thin_border
            if i >= 3:
                cell.number_format = "#,##0"
        # 汇总行高亮
        if row_idx in summary_rows:
            for cell in row:
                cell.fill = summary_fill
                cell.font = summary_font

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    wb.save(output_path)
    return str(output_path)


def main() -> int:
    # 1) 按商品明细查询
    data_by_period: dict[str, dict[str, dict]] = {}
    for key, (start, end) in PERIODS.items():
        print(f"[INFO] querying {key} ...", file=sys.stderr)
        data_by_period[key] = _run_query(PRODUCT_IDS, start, end)
        print(f"[INFO] {key}: {len(data_by_period[key])} products found", file=sys.stderr)

    # 2) 按分类汇总查询
    group_data_by_period: dict[str, dict[str, dict]] = {}
    for key, (start, end) in PERIODS.items():
        group_data_by_period[key] = {}
        for group_name, ids in CATEGORY_GROUPS.items():
            group_data_by_period[key][group_name] = _run_group_query(ids, start, end)
        print(f"[INFO] {key}: {len(group_data_by_period[key])} groups summarized", file=sys.stderr)

    rows: list[list] = []
    summary_rows: list[int] = []

    # 3) 按分类输出：明细行 + 汇总行
    for group_name, ids in CATEGORY_GROUPS.items():
        # 商品明细行
        for pid in ids:
            product_data = {k: v.get(pid, {}) for k, v in data_by_period.items()}
            base = product_data.get("Y26_0506_0621", {})
            rows.append(_build_row(
                base.get("spu_category", ""),
                base.get("spu_product_class", ""),
                pid,
                product_data,
            ))
        # 分类汇总行
        summary_rows.append(len(rows) + 2)  # Excel 行号：header=1
        group_period_data = {k: v.get(group_name, {}) for k, v in group_data_by_period.items()}
        rows.append(_build_row(
            group_name,
            "",
            group_name,
            group_period_data,
        ))

    today = datetime.now()
    base_year = 2026
    date_range = "商品新老客对比"
    out_dir = build_take_path("商品新老客对比", base_year, date_range, extension="xlsx").parent
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = out_dir / "product_new_old_comparison.xlsx"
    csv_path = out_dir / "product_new_old_comparison.csv"

    _write_excel(rows, summary_rows, xlsx_path)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(_make_header())
        writer.writerows(rows)

    print(f"[OK] Excel: {xlsx_path}", file=sys.stderr)
    print(f"[OK] CSV: {csv_path}", file=sys.stderr)
    print(f"\nOutput: {out_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
