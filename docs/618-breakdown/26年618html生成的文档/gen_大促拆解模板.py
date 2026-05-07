#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
618新客预估拆解 Excel 模板生成器
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

OUTPUT_PATH = "/Users/hutou/Desktop/fuqin date/2026_618_新客预估拆解模板_v2.xlsx"

# ============== 样式定义 ==============
def create_styles(wb):
    styles = {}

    # 标题样式
    title = NamedStyle(name="title")
    title.font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="4472C4")
    title.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(title)
    styles["title"] = title

    # 表头样式
    header = NamedStyle(name="header")
    header.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="5B9BD5")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(header)
    styles["header"] = header

    # 普通单元格
    cell = NamedStyle(name="cell")
    cell.font = Font(name="微软雅黑", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(cell)
    styles["cell"] = cell

    # 输入单元格（黄色背景）
    input_cell = NamedStyle(name="input_cell")
    input_cell.font = Font(name="微软雅黑", size=10, bold=True, color="C65911")
    input_cell.fill = PatternFill("solid", fgColor="FFF2CC")
    input_cell.alignment = Alignment(horizontal="center", vertical="center")
    input_cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(input_cell)
    styles["input_cell"] = input_cell

    # 公式单元格（浅绿）
    formula = NamedStyle(name="formula")
    formula.font = Font(name="微软雅黑", size=10, color="006100")
    formula.fill = PatternFill("solid", fgColor="E2EFDA")
    formula.alignment = Alignment(horizontal="center", vertical="center")
    formula.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(formula)
    styles["formula"] = formula

    # 汇总行
    total = NamedStyle(name="total")
    total.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    total.fill = PatternFill("solid", fgColor="70AD47")
    total.alignment = Alignment(horizontal="center", vertical="center")
    total.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    wb.add_named_style(total)
    styles["total"] = total

    # 注释文本
    note = NamedStyle(name="note")
    note.font = Font(name="微软雅黑", size=9, italic=True, color="7F7F7F")
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wb.add_named_style(note)
    styles["note"] = note

    return styles


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============== 参数配置 Sheet ==============
def build_params_sheet(ws, styles):
    ws.title = "参数配置"
    set_col_widths(ws, [20, 25, 50])

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "2026年618大促 新客预估拆解参数配置"
    c.style = styles["title"]
    ws.row_dimensions[1].height = 35

    # 参数列表
    params = [
        ("活动名称", "2026年618大促", "可修改"),
        ("活动开始日期", "2026-05-06", "今年活动开始日期"),
        ("活动结束日期", "2026-06-20", "今年活动结束日期"),
        ("", "", ""),
        ("去年同期开始", "2025-05-06", "用于同比基线"),
        ("去年同期结束", "2025-06-20", "用于同比基线"),
        ("", "", ""),
        ("目标总GSV (万元)", 6000, "【输入】店铺总目标，老客+新客"),
        ("老客占比目标 (%)", 55, "【输入】老客贡献占比，如55%"),
        ("新客增长系数", 1.10, "【输入】新客人数相对去年增长倍数"),
        ("老客复购调整系数", 1.20, "【输入】大促期复购率调整（618=1.20）"),
        ("", "", ""),
        ("新客首单转化率 (%)", 4.0, "【输入】UV到首单转化，默认4%"),
        ("新客平均客单价 (AUS)", 150, "【输入】新客人均贡献GSV"),
        ("老客平均客单价 (AUS)", 217, "【输入】老客人均贡献GSV（去年实际）"),
        ("", "", ""),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M"), "自动"),
    ]

    row = 3
    ws["A3"] = "参数项"
    ws["B3"] = "值"
    ws["C3"] = "说明"
    for col in ["A", "B", "C"]:
        ws[f"{col}3"].style = styles["header"]
    ws.row_dimensions[3].height = 25

    for label, val, note in params:
        if label == "" and val == "" and note == "":
            row += 1
            continue
        ws.cell(row=row, column=1, value=label).style = styles["cell"]
        cell_b = ws.cell(row=row, column=2, value=val)
        cell_c = ws.cell(row=row, column=3, value=note)

        if "【输入】" in str(note):
            cell_b.style = styles["input_cell"]
        else:
            cell_b.style = styles["cell"]
        cell_c.style = styles["note"]

        ws.row_dimensions[row].height = 22
        row += 1

    # 添加数据验证：老客占比 0-100
    dv = DataValidation(type="decimal", operator="between", formula1=0, formula2=100)
    dv.prompt = "请输入0-100之间的数字"
    dv.promptTitle = "老客占比"
    ws.add_data_validation(dv)
    dv.add(f"B{3+8}")


# ============== 老客拆解 Sheet ==============
def build_old_customer_sheet(ws, styles):
    ws.title = "老客拆解"
    set_col_widths(ws, [18, 18, 18, 18, 18, 18, 25])

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "老客拆解：R区间分层 → 预估复购人数 → 预估GSV"
    c.style = styles["title"]
    ws.row_dimensions[1].height = 35

    # 表头
    headers = ["R区间", "当前人数", "去年同区间\n复购人数", "去年复购率", "预估复购率", "预估复购人数", "预估GSV (万元)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.style = styles["header"]
    ws.row_dimensions[3].height = 35

    # 数据（基于去年真实数据）
    # R区间数据：去年同区间复购人数和当前人数
    r_data = [
        ("0-30天", 1854320, 18543, 0.010, "=D4*参数配置!$B$12", "=C4*E4", "=F4*参数配置!$B$15/10000"),
        ("31-90天", 1243200, 6216, 0.005, "=D5*参数配置!$B$12", "=C5*E5", "=F5*参数配置!$B$15/10000"),
        ("91-180天", 876500, 2630, 0.003, "=D6*参数配置!$B$12", "=C6*E6", "=F6*参数配置!$B$15/10000"),
        ("181-365天", 654300, 1309, 0.002, "=D7*参数配置!$B$12", "=C7*E7", "=F7*参数配置!$B$15/10000"),
        ("366天以上", 643237, 643, 0.001, "=D8*参数配置!$B$12", "=C8*E8", "=F8*参数配置!$B$15/10000"),
    ]

    for idx, (r, cur_n, ly_rep, ly_rate, est_rate, est_n, gsv) in enumerate(r_data, 4):
        ws.cell(row=idx, column=1, value=r).style = styles["cell"]
        ws.cell(row=idx, column=2, value=cur_n).style = styles["cell"]
        ws.cell(row=idx, column=3, value=ly_rep).style = styles["cell"]
        ws.cell(row=idx, column=4, value=ly_rate).style = styles["cell"]
        ws.cell(row=idx, column=5, value=est_rate).style = styles["formula"]
        ws.cell(row=idx, column=6, value=est_n).style = styles["formula"]
        ws.cell(row=idx, column=7, value=gsv).style = styles["formula"]
        ws.row_dimensions[idx].height = 25

    # 汇总行
    total_row = 9
    ws.cell(row=total_row, column=1, value="老客合计").style = styles["total"]
    ws.cell(row=total_row, column=2, value="=SUM(B4:B8)").style = styles["total"]
    ws.cell(row=total_row, column=3, value="=SUM(C4:C8)").style = styles["total"]
    ws.cell(row=total_row, column=4, value="=C9/B9").style = styles["total"]
    ws.cell(row=total_row, column=5, value="=F9/B9").style = styles["total"]
    ws.cell(row=total_row, column=6, value="=SUM(F4:F8)").style = styles["total"]
    ws.cell(row=total_row, column=7, value="=SUM(G4:G8)").style = styles["total"]
    ws.row_dimensions[total_row].height = 28

    # 说明
    ws["A11"] = "说明："
    ws["A11"].style = styles["note"]
    notes = [
        "1. 当前人数 = 截至活动开始日（2026-05-06）的老客分层人数",
        "2. 去年同区间复购人数 = 2025-05-06~2025-06-20 该R区间实际复购人数",
        "3. 预估复购率 = 去年复购率 × 老客复购调整系数（618大促1.20）",
        "4. 预估GSV = 预估复购人数 × 老客AUS ÷ 10000（单位：万元）",
        "5. 老客占比 = 老客GSV ÷ 目标总GSV，应与参数配置中的老客占比目标对齐",
    ]
    for i, n in enumerate(notes, 12):
        ws.merge_cells(f"A{i}:G{i}")
        cell = ws.cell(row=i, column=1, value=n)
        cell.style = styles["note"]


# ============== 新客拆解 Sheet ==============
def build_new_customer_sheet(ws, styles):
    ws.title = "新客拆解"
    set_col_widths(ws, [14, 16, 16, 14, 14, 14, 14, 16, 18])

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "新客拆解：渠道UV → 入会率 → 首单转化 → 预估人数 → 预估GSV"
    c.style = styles["title"]
    ws.row_dimensions[1].height = 35

    # 表头
    headers = [
        "渠道", "去年UV", "预估UV", "入会率(%)",
        "首单转化率(%)", "预估新客人数", "AUS(元)", "预估GSV(万元)", "渠道占比"
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.style = styles["header"]
    ws.row_dimensions[3].height = 35

    # 渠道数据（基于去年真实数据）
    # 渠道: [去年UV, 预估UV系数, 入会率, 首单转化率, AUS]
    channel_data = [
        ("货架", 1800000, "=B4*参数配置!$B$11", 2.5, 4.0, 160, "=C4*D4%/100*E4%/100", "=F4*G4/10000", "=H4/H13"),
        ("达播", 500000, "=B5*参数配置!$B$11", 3.0, 4.5, 2435, "=C5*D5%/100*E5%/100", "=F5*G5/10000", "=H5/H13"),
        ("直播", 850000, "=B6*参数配置!$B$11", 2.8, 4.2, 180, "=C6*D6%/100*E6%/100", "=F6*G6/10000", "=H6/H13"),
        ("淘客", 420000, "=B7*参数配置!$B$11", 2.2, 3.8, 140, "=C7*D7%/100*E7%/100", "=F7*G7/10000", "=H7/H13"),
        ("微博", 200000, "=B8*参数配置!$B$11", 2.0, 3.5, 130, "=C8*D8%/100*E8%/100", "=F8*G8/10000", "=H8/H13"),
        ("U先派样", 250000, "=B9*参数配置!$B$11", 15.0, 25.0, 80, "=C9*D9%/100*E9%/100", "=F9*G9/10000", "=H9/H13"),
        ("百补派样", 150000, "=B10*参数配置!$B$11", 12.0, 20.0, 85, "=C10*D10%/100*E10%/100", "=F10*G10/10000", "=H10/H13"),
        ("赠品&0.01", 80000, "=B11*参数配置!$B$11", 18.0, 30.0, 75, "=C11*D11%/100*E11%/100", "=F11*G11/10000", "=H11/H13"),
        ("其他", 200000, "=B12*参数配置!$B$11", 1.5, 2.5, 100, "=C12*D12%/100*E12%/100", "=F12*G12/10000", "=H12/H13"),
    ]

    for idx, (ch, ly_uv, est_uv, join_rate, conv_rate, aus, est_n, gsv, pct) in enumerate(channel_data, 4):
        ws.cell(row=idx, column=1, value=ch).style = styles["cell"]
        ws.cell(row=idx, column=2, value=ly_uv).style = styles["cell"]
        ws.cell(row=idx, column=3, value=est_uv).style = styles["formula"]
        ws.cell(row=idx, column=4, value=join_rate).style = styles["input_cell"]
        ws.cell(row=idx, column=5, value=conv_rate).style = styles["input_cell"]
        ws.cell(row=idx, column=6, value=est_n).style = styles["formula"]
        ws.cell(row=idx, column=7, value=aus).style = styles["input_cell"]
        ws.cell(row=idx, column=8, value=gsv).style = styles["formula"]
        ws.cell(row=idx, column=9, value=pct).style = styles["formula"]
        ws.row_dimensions[idx].height = 25

    # 汇总行
    total_row = 13
    ws.cell(row=total_row, column=1, value="新客合计").style = styles["total"]
    ws.cell(row=total_row, column=2, value="=SUM(B4:B12)").style = styles["total"]
    ws.cell(row=total_row, column=3, value="=SUM(C4:C12)").style = styles["total"]
    ws.cell(row=total_row, column=4, value="=SUMPRODUCT(C4:C12,D4:D12)/C13/100").style = styles["total"]
    ws.cell(row=total_row, column=5, value="=SUMPRODUCT(C4:C12,E4:E12)/C13/100").style = styles["total"]
    ws.cell(row=total_row, column=6, value="=SUM(F4:F12)").style = styles["total"]
    ws.cell(row=total_row, column=7, value="=SUMPRODUCT(F4:F12,G4:G12)/F13").style = styles["total"]
    ws.cell(row=total_row, column=8, value="=SUM(H4:H12)").style = styles["total"]
    ws.cell(row=total_row, column=9, value="=SUM(I4:I12)").style = styles["total"]
    ws.row_dimensions[total_row].height = 28

    # 说明
    ws["A15"] = "说明："
    ws["A15"].style = styles["note"]
    notes = [
        "1. 预估UV = 去年UV × 新客增长系数（参数配置中设置，默认1.10）",
        "2. 入会率和首单转化率【黄色区域可手动调整】，不同渠道差异很大",
        "3. 预估新客人数 = 预估UV × 入会率% × 首单转化率%",
        "4. 预估GSV = 预估新客人数 × AUS ÷ 10000（单位：万元）",
        "5. 派样渠道（U先/百补/赠品）入会率和转化率显著高于常规渠道",
        "6. 渠道占比 = 该渠道GSV ÷ 新客总GSV",
    ]
    for i, n in enumerate(notes, 16):
        ws.merge_cells(f"A{i}:I{i}")
        cell = ws.cell(row=i, column=1, value=n)
        cell.style = styles["note"]


# ============== 汇总 Sheet ==============
def build_summary_sheet(ws, styles):
    ws.title = "汇总"
    set_col_widths(ws, [22, 20, 20, 20, 38])

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "2026年618大促 新老客预估汇总"
    c.style = styles["title"]
    ws.row_dimensions[1].height = 35

    # 汇总表
    headers = ["项目", "预估人数", "预估GSV (万元)", "占比", "备注"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.style = styles["header"]
    ws.row_dimensions[3].height = 28

    # 老客行
    ws.cell(row=4, column=1, value="老客复购").style = styles["cell"]
    ws.cell(row=4, column=2, value="=老客拆解!F9").style = styles["formula"]
    ws.cell(row=4, column=3, value="=老客拆解!G9").style = styles["formula"]
    ws.cell(row=4, column=4, value="=C4/C6").style = styles["formula"]
    ws.cell(row=4, column=5, value="基于R区间分层复购预估").style = styles["note"]
    ws.row_dimensions[4].height = 25

    # 新客行
    ws.cell(row=5, column=1, value="新客首购").style = styles["cell"]
    ws.cell(row=5, column=2, value="=新客拆解!F13").style = styles["formula"]
    ws.cell(row=5, column=3, value="=新客拆解!H13").style = styles["formula"]
    ws.cell(row=5, column=4, value="=C5/C6").style = styles["formula"]
    ws.cell(row=5, column=5, value="基于渠道UV×转化预估").style = styles["note"]
    ws.row_dimensions[5].height = 25

    # 合计行
    ws.cell(row=6, column=1, value="合计").style = styles["total"]
    ws.cell(row=6, column=2, value="=B4+B5").style = styles["total"]
    ws.cell(row=6, column=3, value="=C4+C5").style = styles["total"]
    ws.cell(row=6, column=4, value="=C6/参数配置!$B$8").style = styles["total"]
    ws.cell(row=6, column=5, value="对比目标总GSV").style = styles["total"]
    ws.row_dimensions[6].height = 28

    # 目标对比
    ws.cell(row=8, column=1, value="目标总GSV").style = styles["cell"]
    ws.cell(row=8, column=2, value="-").style = styles["cell"]
    ws.cell(row=8, column=3, value="=参数配置!$B$8").style = styles["cell"]
    ws.cell(row=8, column=4, value="100%").style = styles["cell"]
    ws.cell(row=8, column=5, value="参数配置中设定").style = styles["note"]
    ws.row_dimensions[8].height = 25

    ws.cell(row=9, column=1, value="缺口/盈余").style = styles["cell"]
    ws.cell(row=9, column=2, value="-").style = styles["cell"]
    ws.cell(row=9, column=3, value="=C6-C8").style = styles["formula"]
    ws.cell(row=9, column=4, value="=C9/C8").style = styles["formula"]
    ws.cell(row=9, column=5, value="正值=超额完成，负值=缺口").style = styles["note"]
    ws.row_dimensions[9].height = 25

    # 关键指标对比
    ws.merge_cells("A11:E11")
    ws["A11"] = "关键指标对比"
    ws["A11"].style = styles["header"]
    ws.row_dimensions[11].height = 28

    comp_headers = ["指标", "去年同期", "今年预估", "变化", "说明"]
    for i, h in enumerate(comp_headers, 1):
        cell = ws.cell(row=12, column=i, value=h)
        cell.style = styles["header"]

    comp_data = [
        ("总UV", 4170000, "=新客拆解!C13", "=C13/B13-1", " UV总量"),
        ("总新客人数", 138341, "=新客拆解!F13", "=C14/B14-1", " 首购人数"),
        ("总GSV (万元)", 3955, "=C6", "=C15/B15-1", " 老客+新客"),
        ("客单价 (元)", 160, "=C6*10000/B6", "=C16/B16-1", " 人均GSV"),
    ]
    for idx, (label, ly, cur, change, note) in enumerate(comp_data, 13):
        ws.cell(row=idx, column=1, value=label).style = styles["cell"]
        ws.cell(row=idx, column=2, value=ly).style = styles["cell"]
        ws.cell(row=idx, column=3, value=cur).style = styles["formula"]
        ws.cell(row=idx, column=4, value=change).style = styles["formula"]
        ws.cell(row=idx, column=5, value=note).style = styles["note"]
        ws.row_dimensions[idx].height = 25

    # 底部说明
    ws.merge_cells("A18:E18")
    ws["A18"] = "使用说明：在【参数配置】Sheet中修改黄色区域的参数，其余Sheet会自动联动计算。"
    ws["A18"].style = styles["note"]
    ws.row_dimensions[18].height = 30


# ============== 主程序 ==============
def main():
    wb = openpyxl.Workbook()
    styles = create_styles(wb)

    # 删除默认sheet
    wb.remove(wb.active)

    # 按顺序创建
    ws_params = wb.create_sheet("参数配置", 0)
    ws_old = wb.create_sheet("老客拆解", 1)
    ws_new = wb.create_sheet("新客拆解", 2)
    ws_summary = wb.create_sheet("汇总", 3)

    build_params_sheet(ws_params, styles)
    build_old_customer_sheet(ws_old, styles)
    build_new_customer_sheet(ws_new, styles)
    build_summary_sheet(ws_summary, styles)

    # 冻结窗格
    ws_params.freeze_panes = "A4"
    ws_old.freeze_panes = "A4"
    ws_new.freeze_panes = "A4"
    ws_summary.freeze_panes = "A4"

    wb.save(OUTPUT_PATH)
    print(f"模板已生成: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
