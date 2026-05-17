"""
话术库服务 - 产品话术管理
"""
import json
from pathlib import Path
from typing import List, Optional, Dict, Any
from datetime import datetime

# 话术库存储路径
SCRIPT_LIBRARY_DIR = Path(__file__).parent.parent.parent / "data" / "script_library"
PRODUCT_SCRIPTS_FILE = SCRIPT_LIBRARY_DIR / "product_scripts.json"
CELEBRITY_SCRIPTS_FILE = SCRIPT_LIBRARY_DIR / "celebrity_scripts.json"


def _ensure_dir():
    """确保目录存在"""
    SCRIPT_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)


def _load_json(filepath: Path) -> Dict[str, Any]:
    """加载JSON文件"""
    if filepath.exists():
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"version": "1.0.0", "updated_at": None, "products": {}}


def _save_json(filepath: Path, data: Dict[str, Any]):
    """保存JSON文件"""
    _ensure_dir()
    data["updated_at"] = datetime.now().isoformat()
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ==================== 产品话术 ====================

def get_product_list() -> List[Dict[str, Any]]:
    """获取产品列表"""
    data = _load_json(PRODUCT_SCRIPTS_FILE)
    products = data.get("products", {})
    return [
        {
            "name": name,
            "qa_count": len(info.get("qa_list", [])),
            "updated_at": info.get("updated_at")
        }
        for name, info in products.items()
    ]


def get_product_scripts(product_name: str) -> Optional[Dict[str, Any]]:
    """获取指定产品的所有话术"""
    data = _load_json(PRODUCT_SCRIPTS_FILE)
    products = data.get("products", {})
    return products.get(product_name)


def search_product_scripts(keyword: str) -> List[Dict[str, Any]]:
    """搜索产品话术（按问题关键词）"""
    data = _load_json(PRODUCT_SCRIPTS_FILE)
    results = []
    for product_name, info in data.get("products", {}).items():
        for qa in info.get("qa_list", []):
            if keyword.lower() in qa.get("question", "").lower():
                results.append({
                    "product": product_name,
                    "question": qa["question"],
                    "answer": qa.get("answer", "")
                })
    return results


def import_product_scripts(excel_path: str) -> Dict[str, Any]:
    """从Excel导入产品话术"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    data = _load_json(PRODUCT_SCRIPTS_FILE)
    products = data.setdefault("products", {})

    imported_count = 0
    skipped_count = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == "目录":
            continue

        ws = wb[sheet_name]
        current_product = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                continue

            col_a, col_b, col_c, col_d = row[0], row[1], row[2], row[3]

            # 更新产品名
            if col_a and str(col_a).strip():
                raw_name = str(col_a).strip().split('\n')[0].strip()
                if raw_name not in {"产品", "通用", "售后问题", "共性问题", "全部款"}:
                    current_product = raw_name

            if not current_product or not col_b:
                continue

            question = str(col_b).strip()
            answer = str(col_d).strip() if col_d else ""

            # 初始化产品
            if current_product not in products:
                products[current_product] = {"qa_list": [], "updated_at": datetime.now().isoformat()}

            # 检查是否已存在
            existing_questions = {qa["question"] for qa in products[current_product]["qa_list"]}
            if question not in existing_questions:
                products[current_product]["qa_list"].append({
                    "question": question,
                    "answer": answer,
                    "has_answer": bool(answer and answer != "None")
                })
                imported_count += 1
            else:
                skipped_count += 1

    _save_json(PRODUCT_SCRIPTS_FILE, data)

    return {
        "imported": imported_count,
        "skipped": skipped_count,
        "total_products": len(products)
    }


# ==================== 明星专项 ====================

# 预设的明星数据
DEFAULT_CELEBRITIES = [
    {
        "name": "鞠婧祎",
        "avatar": "https://img.example.com/ju.jpg",
        "title": "品牌代言人",
        "status": "待开发",
        "scripts": []
    },
    {
        "name": "王一博",
        "avatar": "https://img.example.com/wang.jpg",
        "title": "品牌代言人",
        "status": "待开发",
        "scripts": []
    },
    {
        "name": "王鹤棣",
        "avatar": "https://img.example.com/wanghd.jpg",
        "title": "品牌代言人",
        "status": "待开发",
        "scripts": []
    },
    {
        "name": "张凌赫",
        "avatar": "https://img.example.com/zhang.jpg",
        "title": "品牌代言人",
        "status": "待开发",
        "scripts": []
    }
]


def get_celebrity_list() -> List[Dict[str, Any]]:
    """获取明星列表"""
    data = _load_json(CELEBRITY_SCRIPTS_FILE)
    celebrities = data.get("celebrities", [])

    # 如果没有数据，初始化默认数据
    if not celebrities:
        data["celebrities"] = DEFAULT_CELEBRITIES
        _save_json(CELEBRITY_SCRIPTS_FILE, data)
        celebrities = DEFAULT_CELEBRITIES

    return celebrities


def get_celebrity_scripts(celebrity_name: str) -> Optional[Dict[str, Any]]:
    """获取指定明星的话术"""
    data = _load_json(CELEBRITY_SCRIPTS_FILE)
    for celeb in data.get("celebrities", []):
        if celeb["name"] == celebrity_name:
            return celeb
    return None


def update_celebrity_scripts(celebrity_name: str, scripts: List[Dict[str, str]]) -> bool:
    """更新明星话术"""
    data = _load_json(CELEBRITY_SCRIPTS_FILE)
    celebrities = data.get("celebrities", [])

    for celeb in celebrities:
        if celeb["name"] == celebrity_name:
            celeb["scripts"] = scripts
            celeb["status"] = "已上线" if scripts else "待开发"
            _save_json(CELEBRITY_SCRIPTS_FILE, data)
            return True

    return False


# ==================== 初始化示例数据 ====================

def init_sample_data():
    """初始化示例数据（仅在文件不存在时）"""
    _ensure_dir()

    # 产品话术示例
    if not PRODUCT_SCRIPTS_FILE.exists():
        sample_products = {
            "B5面膜": {
                "qa_list": [
                    {"question": "B5面膜成分有哪些", "answer": "B5面膜仅含9种成分，核心成分是3%泛醇+5%积雪草+1%油橄榄", "has_answer": True},
                    {"question": "敏感肌可以用吗", "answer": "可以的，B5面膜专为敏感肌设计，成分精简温和", "has_answer": True},
                    {"question": "多久用一次", "answer": "建议每周2-3次，每次15-20分钟", "has_answer": True}
                ],
                "updated_at": datetime.now().isoformat()
            },
            "09水乳": {
                "qa_list": [
                    {"question": "09水乳适合什么肤质", "answer": "适合干敏/油敏肌肤，有滋润版和清爽版", "has_answer": True},
                    {"question": "成分有多少种", "answer": "每款仅9种成分，从8000多种原料中筛选", "has_answer": True}
                ],
                "updated_at": datetime.now().isoformat()
            }
        }
        _save_json(PRODUCT_SCRIPTS_FILE, {"products": sample_products, "version": "1.0.0"})

    # 明星专项示例
    if not CELEBRITY_SCRIPTS_FILE.exists():
        _save_json(CELEBRITY_SCRIPTS_FILE, {"celebrities": DEFAULT_CELEBRITIES, "version": "1.0.0"})


# 启动时初始化
init_sample_data()
